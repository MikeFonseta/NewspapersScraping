import openpyxl
from webdriver_manager.microsoft import EdgeChromiumDriverManager
from selenium.webdriver.edge.service import Service
from selenium import webdriver
from selenium.webdriver.edge.options import Options as EdgeOptions
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
from selenium.webdriver.common.by import By

#Opzioni per il webdriver
options = EdgeOptions()
options.add_argument("start-maximized")
options.add_experimental_option('excludeSwitches', ['enable-logging'])

#Creazione webdriver           In questo caso Edge
driver = webdriver.Edge(service=Service(EdgeChromiumDriverManager().install()),options=options)

nomeFile = 'BresciaOggi'

wb = openpyxl.load_workbook(nomeFile+'.xlsx')
 
sheet = wb.active
rows = sheet.max_row

#For che scorre tutte le righe del foglio excel e prende il valore della terza colonna
for i in range(1, rows + 1):
    cell = sheet.cell(row = i, column = 3)
    cellToWrite = sheet.cell(row = i, column = 4)

    #caricamento pagina
    driver.get(cell.value)
    
    try:
        cellToWrite.value = WebDriverWait(driver, 10).until(EC.visibility_of_element_located((By.CLASS_NAME, 'content-body'))).text
        wb.save(nomeFile+'.xlsx')
    except:
        cellToWrite.value = "Error"

wb.save(nomeFile+'.xlsx')
print("Terminato")
